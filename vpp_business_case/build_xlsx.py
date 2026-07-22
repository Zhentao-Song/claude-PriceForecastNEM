# -*- coding: utf-8 -*-
"""Polished VPP business-case workbook (StarCharge style)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

OUT="/sessions/funny-upbeat-darwin/mnt/claude-PriceForecastNEM/vpp_business_case/VPP收益测算表.xlsx"
ORANGE="EC6A2C"; TINT="FCEEE3"; GREY="F3F3F3"; DARK="2E2E2E"; HEADTXT="FFFFFF"
BLUE="0000FF"; MUTE="7A7A7A"
FNAME="Arial"

thin=Side(style="thin",color="D9D9D9")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=10,b=False,c=DARK): return Font(name=FNAME,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
right=Alignment(horizontal="right",vertical="center")

wb=Workbook()

def style_header(ws,row,c1,c2,labels,widths=None):
    for j,lab in enumerate(labels):
        cell=ws.cell(row=row,column=c1+j,value=lab)
        cell.fill=fill(ORANGE); cell.font=F(10.5,True,HEADTXT); cell.alignment=center; cell.border=border
    if widths:
        for j,w in enumerate(widths): ws.column_dimensions[get_column_letter(c1+j)].width=w

def title_bar(ws,text,span,row=1):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row=row,column=1,value=text)
    c.fill=fill(ORANGE); c.font=F(15,True,HEADTXT); c.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[row].height=30

# ============== Sheet 1: 测算总览 ==============
ws=wb.active; ws.title="测算总览"
ws.sheet_view.showGridLines=False
title_bar(ws,"星星充电 · VPP收益测算总览 (AUD, 基于AEMO NEM NSW1真实数据)",8)
ws.cell(row=2,column=1,value="数据基准：AEMO DISPATCHPRICE 真实5分钟数据 (NSW1, 2025 Q1, 含RRP+8个FCAS服务价)  ·  口径保守：季节0.80×预见性0.85×单循环").font=F(9,False,MUTE)

# headline metrics
ws.cell(row=4,column=1,value="里程碑收益（运营商口径）").font=F(12,True,ORANGE)
hm_hdr_row=5
style_header(ws,hm_hdr_row,1,4,["阶段","聚合规模","运营商收益(万AUD/年)","说明"],[16,14,22,30])
hm=[("首批 · 试点","50MW","='聚合规模与收益'!H7","2026 Q4 上线"),
    ("中期 · 储备","320MW","='聚合规模与收益'!H11","2027–2028"),
    ("远期 · 规划","650MW","='聚合规模与收益'!H15","2029+ · IRP独立持牌")]
r=hm_hdr_row+1
for name,scale,formula,note in hm:
    ws.cell(row=r,column=1,value=name).font=F(10.5,True)
    ws.cell(row=r,column=2,value=scale).font=F(10.5)
    cc=ws.cell(row=r,column=3,value=formula); cc.font=F(12,True,ORANGE)
    cc.number_format='#,##0.0'
    ws.cell(row=r,column=4,value=note).font=F(10,False,MUTE)
    for col in range(1,5):
        ws.cell(row=r,column=col).border=border; ws.cell(row=r,column=col).alignment=center
    ws.cell(row=r,column=4).alignment=left
    r+=1

# assumptions
ws.cell(row=11,column=1,value="关键假设（蓝色=可调输入）").font=F(12,True,ORANGE)
style_header(ws,12,1,2,["参数","数值"])
A=[("运营商 VPP 分成",0.30,'0%'),
   ("BESS 运营分成（自持兜底）",1.00,'0%'),
   ("电池往返效率",0.90,'0%'),
   ("季节折减 (夏季Q1→全年)",0.80,'0.00'),
   ("预见性折减 (尖峰捕获)",0.85,'0.00'),
   ("BESS 运营单位收益 (AUD/MW·年)",40000,'#,##0'),
   ("BESS 现货套利价差 (AUD/MWh)",190,'#,##0')]
r=13
for lab,val,fmt in A:
    ws.cell(row=r,column=1,value=lab).font=F(10); ws.cell(row=r,column=1).border=border; ws.cell(row=r,column=1).alignment=left
    c=ws.cell(row=r,column=2,value=val); c.font=F(10,False,BLUE); c.number_format=fmt; c.border=border; c.alignment=center
    r+=1
# named refs used elsewhere: 分成=B13, BESS分成=B14, BESS单位=B18
ws.cell(row=21,column=1,value="图例：蓝色=硬输入  ·  黑色=公式计算  ·  绿色=跨表引用").font=F(9,False,MUTE)

# ============== Sheet 2: 单位经济性测算 ==============
ws2=wb.create_sheet("单位经济性测算")
ws2.sheet_view.showGridLines=False
title_bar(ws2,"单位经济性测算（每户 / 每站，AUD/年）",4)
ws2.cell(row=2,column=1,value="方案A=零售套餐·无VPP；方案B=批发套餐·VPP(套利+FCAS)。蓝色为NEM真实回测输出。").font=F(9,False,MUTE)
style_header(ws2,3,1,3,["指标","户用（每户）","工商业（每站）"],[34,18,18])
rows=[
 ("储能容量 (kWh)",13.5,215,'#,##0.0','in'),
 ("功率 (kW)",5,100,'#,##0','in'),
 ("方案A 零售套利 VPP收益",931,8283,'#,##0','in'),
 ("方案B 批发套利",904,11480,'#,##0','in'),
 ("方案B FCAS",179,3810,'#,##0','in'),
 ("方案B VPP收益合计","=B7+B8","=C7+C8",'#,##0','f'),
 ("提升倍数 (B/A)","=B9/B6","=C9/C6",'0.00"×"','f'),
 ("单位VPP收益 (AUD/kW·年)","=B9/B5","=C9/C5",'#,##0.0','f'),
 ("运营商分成30% (AUD/年)","=B9*测算总览!$B$13","=C9*测算总览!$B$13",'#,##0','f'),
 ("运营商单位收益 (AUD/kW·年)","=B11*测算总览!$B$13","=C11*测算总览!$B$13",'#,##0.0','f'),
]
r=4
for lab,bv,cv,fmt,kind in rows:
    ws2.cell(row=r,column=1,value=lab).font=F(10,kind=='f');
    ws2.cell(row=r,column=1).alignment=left; ws2.cell(row=r,column=1).border=border
    for col,val in ((2,bv),(3,cv)):
        c=ws2.cell(row=r,column=col,value=val)
        c.number_format=fmt; c.border=border; c.alignment=center
        c.font=F(10,kind=='f', BLUE if kind=='in' else DARK)
    # highlight total + operator rows
    if lab.startswith("方案B VPP收益合计") or lab.startswith("运营商分成"):
        for col in range(1,4):
            ws2.cell(row=r,column=col).fill=fill(TINT); ws2.cell(row=r,column=col).font=F(10,True, ORANGE if col>1 else DARK)
    r+=1

# ============== Sheet 3: 聚合规模与收益 ==============
ws3=wb.create_sheet("聚合规模与收益")
ws3.sheet_view.showGridLines=False
title_bar(ws3,"聚合规模与收益（按第1页口径，万 AUD/年）",9)
hdr=["阶段","资产类型","聚合规模(MW)","单位收益","单位口径","毛收益(万AUD/年)","分成","运营商收益(万AUD/年)","对标数据来源"]
style_header(ws3,3,1,9,hdr,[12,13,13,11,13,16,8,18,24])
# data: stage, asset, MW, unit(formula/val), unitlabel, gross_formula, share_ref, src
def seg_rows(stage, resi, ci, bess):
    return [
     (stage,"户用 VPP",resi,"=单位经济性测算!$B$11","AUD/kW·年","VPP","=测算总览!$B$13","AEMO NEM 2024–25"),
     ("","工商业 VPP",ci,"=单位经济性测算!$C$11","AUD/kW·年","VPP","=测算总览!$B$13","Momentum / EnelX 参考"),
     ("","BESS 运营",bess,"=测算总览!$B$18","AUD/MW·年","BESS","=测算总览!$B$14","Modo Energy 2025 BESS"),
    ]
blocks=[("首批 · 试点",10,20,20),("中期 · 储备",20,50,250),("远期 · 规划",50,100,500)]
r=4
subtotal_rows=[]
for stage,resi,ci,bess in blocks:
    start=r
    for st,asset,mw,unit,unitlab,kind,share,src in seg_rows(stage,resi,ci,bess):
        ws3.cell(row=r,column=1,value=st).font=F(10,True)
        ws3.cell(row=r,column=2,value=asset).font=F(10)
        ws3.cell(row=r,column=3,value=mw).font=F(10,False,BLUE); ws3.cell(row=r,column=3).number_format='#,##0" MW"'
        ws3.cell(row=r,column=4,value=unit).font=F(10); ws3.cell(row=r,column=4).number_format='#,##0.#'
        ws3.cell(row=r,column=5,value=unitlab).font=F(9,False,MUTE)
        # gross: VPP =C*D/10 ; BESS =C*D/10000
        if kind=="VPP":
            ws3.cell(row=r,column=6,value=f"=C{r}*D{r}/10")
        else:
            ws3.cell(row=r,column=6,value=f"=C{r}*D{r}/10000")
        ws3.cell(row=r,column=6).number_format='#,##0.0'; ws3.cell(row=r,column=6).font=F(10)
        ws3.cell(row=r,column=7,value=share); ws3.cell(row=r,column=7).number_format='0%'; ws3.cell(row=r,column=7).font=F(10)
        ws3.cell(row=r,column=8,value=f"=F{r}*G{r}"); ws3.cell(row=r,column=8).number_format='#,##0.0'; ws3.cell(row=r,column=8).font=F(10,True)
        ws3.cell(row=r,column=9,value=src).font=F(9,False,MUTE)
        for col in range(1,10):
            ws3.cell(row=r,column=col).border=border
            ws3.cell(row=r,column=col).alignment= left if col in (5,9) else center
        r+=1
    # subtotal
    ws3.cell(row=r,column=1,value="小计").font=F(10,True,ORANGE)
    ws3.cell(row=r,column=2,value=stage.split(" ")[0]).font=F(10,True,ORANGE)
    ws3.cell(row=r,column=3,value=f"=SUM(C{start}:C{r-1})").number_format='#,##0" MW"'
    ws3.cell(row=r,column=6,value=f"=SUM(F{start}:F{r-1})").number_format='#,##0.0'
    ws3.cell(row=r,column=8,value=f"=SUM(H{start}:H{r-1})").number_format='#,##0.0'
    for col in range(1,10):
        cc=ws3.cell(row=r,column=col); cc.fill=fill(TINT); cc.border=border
        if cc.font.color is None or col in (3,6,8): cc.font=F(10.5,True,ORANGE if col in (6,8) else DARK)
        cc.alignment=center
    subtotal_rows.append(r)
    r+=1

# ============== Sheet 4: 增长轨迹 ==============
ws4=wb.create_sheet("增长轨迹")
ws4.sheet_view.showGridLines=False
title_bar(ws4,"JV 收入增长轨迹（运营商收益，万 AUD/年）",5)
style_header(ws4,3,1,5,["时点","民居 VPP","工商业 VPP","BESS 运营","合计"],[14,14,14,14,12])
traj=[("2026 Q4",65,92,80),("2027",100,160,450),("2028 H1",130,230,1000),("2029+",324,459,2000)]
r=4
for t,a,b,c in traj:
    ws4.cell(row=r,column=1,value=t).font=F(10,True)
    for col,val in ((2,a),(3,b),(4,c)):
        cc=ws4.cell(row=r,column=col,value=val); cc.font=F(10,False,BLUE); cc.number_format='#,##0'
    ws4.cell(row=r,column=5,value=f"=SUM(B{r}:D{r})").font=F(10,True,ORANGE); ws4.cell(row=r,column=5).number_format='#,##0'
    for col in range(1,6): ws4.cell(row=r,column=col).border=border; ws4.cell(row=r,column=col).alignment=center
    r+=1
# chart
chart=LineChart(); chart.title="JV 收入增长轨迹 (万 AUD/年)"; chart.style=2
chart.height=8.5; chart.width=16
data=Reference(ws4,min_col=2,max_col=4,min_row=3,max_row=7)
cats=Reference(ws4,min_col=1,min_row=4,max_row=7)
chart.add_data(data,titles_from_data=True); chart.set_categories(cats)
chart.y_axis.title="万 AUD/年"; chart.x_axis.title="时点"
ws4.add_chart(chart,"A10")

# ============== Sheet 5: 典型日演示 ==============
ws5=wb.create_sheet("典型日演示")
ws5.sheet_view.showGridLines=False
title_bar(ws5,"典型日演示 · 两场景一天的电池调度与收益（NSW全季平均日内价格曲线）",8)
ws5.cell(row=2,column=1,value="演示口径：按NSW现货全季平均日内曲线，单户/单站一天的最优低充高放(单循环)+FCAS。工商业工作日08–17点电池服务负荷、不参与套利。").font=F(9,False,MUTE)

CHGFILL="DCE6F1"; DISFILL="FCEEE3"; BLKFILL="EDEDED"
PRICES={0:90.2,1:86.1,2:83.9,3:84.8,4:91.0,5:107.2,6:90.4,7:44.1,8:36.1,9:27.6,
        10:21.6,11:23.1,12:20.5,13:40.3,14:109.4,15:67.0,16:98.3,17:205.9,18:211.6,
        19:161.6,20:112.7,21:104.6,22:101.8,23:95.8}
RESI={10:-5,11:-5,12:-5,17:5,18:5,19:3.5}
CI={7:-100,2:-100,3:-38.9,17:100,18:100,19:15}

# summary block (refs to totals below)
ws5.cell(row=4,column=1,value="日收益汇总").font=F(12,True,ORANGE)
style_header(ws5,5,1,6,["场景","套利($/日)","FCAS($/日)","日VPP收益($)","运营商30%($/日)","年化≈($/年)"],
             [16,13,13,15,17,15])
sm=[("户用（每户·13.5kWh/5kW）","E","6"),("工商业（每站·215kWh/100kW）","H","7")]
for name,col,r in [("户用（每户·13.5kWh/5kW）",("E","D"),6),("工商业（每站·215kWh/100kW）",("H","G"),7)]:
    cc=col[0]
    ws5.cell(row=r,column=1,value=name).font=F(10,True); ws5.cell(row=r,column=1).alignment=left
    ws5.cell(row=r,column=2,value=f"={cc}34").number_format='#,##0.00'
    ws5.cell(row=r,column=3,value=f"={cc}35").number_format='#,##0.00'
    ws5.cell(row=r,column=4,value=f"={cc}36").number_format='#,##0.00'
    ws5.cell(row=r,column=5,value=f"={cc}37").number_format='#,##0.00'
    ws5.cell(row=r,column=6,value=f"={cc}38").number_format='#,##0'
    for j in range(1,7):
        cell=ws5.cell(row=r,column=j); cell.border=border
        if j>1: cell.alignment=center; cell.font=F(10,True, ORANGE if j in (4,5) else DARK)
    ws5.cell(row=r,column=1).fill=fill(TINT)

# hourly table
hdr2=["小时","现货价($/MWh)","户用·动作","户用·电量(kWh)","户用·收益($)","工商业·动作","工商业·电量(kWh)","工商业·收益($)"]
style_header(ws5,9,1,8,hdr2,[8,15,12,15,13,14,17,13])
for h in range(24):
    r=10+h
    ws5.cell(row=r,column=1,value=f"{h:02d}:00").font=F(10); ws5.cell(row=r,column=1).alignment=center
    ws5.cell(row=r,column=2,value=PRICES[h]).font=F(10); ws5.cell(row=r,column=2).number_format='#,##0.0'; ws5.cell(row=r,column=2).alignment=center
    # residential
    re_e=RESI.get(h,0)
    ract="充电" if re_e<0 else ("放电" if re_e>0 else "—")
    ws5.cell(row=r,column=3,value=ract); ws5.cell(row=r,column=4,value=re_e if re_e else 0)
    ws5.cell(row=r,column=5,value=f"=B{r}*D{r}/1000")
    # C&I
    if 8<=h<=16:
        cact="服务负荷"; ce_e=0
    else:
        ce_e=CI.get(h,0); cact="充电" if ce_e<0 else ("放电" if ce_e>0 else "—")
    ws5.cell(row=r,column=6,value=cact); ws5.cell(row=r,column=7,value=ce_e if ce_e else 0)
    ws5.cell(row=r,column=8,value=f"=B{r}*G{r}/1000")
    for j in (4,5,7,8): ws5.cell(row=r,column=j).number_format='#,##0.00'
    ws5.cell(row=r,column=4).font=F(10,False,BLUE); ws5.cell(row=r,column=7).font=F(10,False,BLUE)
    for j in range(1,9):
        cell=ws5.cell(row=r,column=j); cell.border=border
        if j in (3,6): cell.alignment=center
        if cell.alignment is None or j in (5,8): cell.alignment=center
    # row tints
    if re_e<0:
        for j in (3,4): ws5.cell(row=r,column=j).fill=fill(CHGFILL)
    elif re_e>0:
        for j in (3,4): ws5.cell(row=r,column=j).fill=fill(DISFILL)
    if 8<=h<=16:
        for j in (6,7): ws5.cell(row=r,column=j).fill=fill(BLKFILL)
    elif ce_e<0:
        for j in (6,7): ws5.cell(row=r,column=j).fill=fill(CHGFILL)
    elif ce_e>0:
        for j in (6,7): ws5.cell(row=r,column=j).fill=fill(DISFILL)

# totals rows 34-38 (data occupies rows 10-33)
def trow(r,label,e,h,fmt='#,##0.00',key=False):
    ws5.cell(row=r,column=1,value=label); ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    ws5.cell(row=r,column=1).alignment=right
    ws5.cell(row=r,column=1).font=F(11,True,ORANGE) if key else F(10.5,True)
    ws5.cell(row=r,column=5,value=e); ws5.cell(row=r,column=8,value=h)
    for j in (5,8):
        c=ws5.cell(row=r,column=j); c.number_format=fmt; c.alignment=center
        c.font=F(11,True,ORANGE) if key else F(10.5,True)
    for j in range(1,9):
        ws5.cell(row=r,column=j).border=border
        if key: ws5.cell(row=r,column=j).fill=fill(TINT)
trow(34,"① 套利合计 / 日","=SUM(E10:E33)","=SUM(H10:H33)")
trow(35,"② FCAS / 日（年度模型÷(365×0.9)）","=单位经济性测算!$B$8/(365*0.9)","=单位经济性测算!$C$8/(365*0.9)")
trow(36,"③ 日VPP收益 = ①+②","=E34+E35","=H34+H35",key=True)
trow(37,"④ 运营商分成30% / 日","=E36*测算总览!$B$13","=H36*测算总览!$B$13")
trow(38,"⑤ 年化≈ 日VPP×365（对照年度模型：户用1,083 / 工商业15,290）","=E36*365","=H36*365",fmt='#,##0')
ws5.cell(row=38,column=1).font=F(10,False,MUTE)
ws5.cell(row=40,column=1,value="说明：电量正=放电入网、负=充电(含效率损耗，户用充15售13.5；工商充238.9售215)。蓝=调度输入，橙=关键结果。典型日略低于年化/365，因平均曲线平滑了真实尖峰，年度模型另计季节/预见性折减。").font=F(9,False,MUTE)
ws5.freeze_panes="A10"

# row heights + page setup
for ws_ in (ws,ws2,ws3,ws4,ws5):
    ws_.row_dimensions[1].height=28
from openpyxl.worksheet.properties import PageSetupProperties
for ws_ in (ws,ws3,ws4,ws5):
    ws_.page_setup.orientation="landscape"
    ws_.page_setup.fitToWidth=1; ws_.page_setup.fitToHeight=0
    ws_.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)

wb.save(OUT)
print("saved", OUT)
