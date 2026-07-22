# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

OUT="/sessions/funny-upbeat-darwin/mnt/claude-PriceForecastNEM/vpp_business_case/VPP竞品对比.xlsx"
ORANGE="EC6A2C"; TINT="FCEEE3"; GREY="F3F3F3"; INK="2E2E2E"; WHITE="FFFFFF"; MUTE="6E6E6E"
F="PingFang SC"
thin=Side(style="thin",color="DADADA"); border=Border(thin,thin,thin,thin)
def FN(sz=10,b=False,c=INK): return Font(name=F,size=sz,bold=b,color=c)
wrap=Alignment(horizontal="left",vertical="center",wrap_text=True)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)

wb=Workbook(); ws=wb.active; ws.title="竞品对比"
ws.sheet_view.showGridLines=False

cols=["类别","名称","平台/产品","面向场景","规模 / 实绩","主要服务","收费模式"]
widths=[12,15,16,17,30,32,24]
# title
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
t=ws.cell(row=1,column=1,value="澳洲 VPP 竞品 / 参考案例对比（聚焦 C&I）")
t.fill=PatternFill("solid",fgColor=ORANGE); t.font=FN(15,True,WHITE)
t.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=30
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(cols))
ws.cell(row=2,column=1,value="数据来源：各公司官网/ARENA/pv magazine/Energy-Storage.News 等公开资料（2024–2026）；规模为公开披露口径").font=FN(8.5,False,MUTE)

# header
for j,c in enumerate(cols,1):
    cell=ws.cell(row=3,column=j,value=c); cell.fill=PatternFill("solid",fgColor=INK)
    cell.font=FN(10.5,True,WHITE); cell.alignment=ctr; cell.border=border
    ws.column_dimensions[get_column_letter(j)].width=widths[j-1]

rows=[
 ("C&I 负荷型\nDR / VPP","Enel X","DER.OS 优化软件","C&I 需求响应 / VPP",
  "澳洲 95MW VPP（21 客户 / 108 站点）；组合 100+MW；全球 DR 10GW",
  "需求响应聚合（WDRM 唯一注册）、表后电池+EV、工业负荷灵活","收益分成 + 可用/激活付费"),
 ("C&I 负荷型\nDR / VPP","Flow Power","kWatch 智能控制器","C&I 现货敞口 + DR",
  "DR 20MW / 95 站点；SA 20.7MW / 43 站点（单点 28kW–2.3MW）",
  "C&I 零售（现货敞口）、负荷 DR、RERT 储备、kWatch 控制","零售合同 + DR 可用/激活付费"),
 ("C&I 负荷型\nDR / VPP","Momentum Energy","自营（背靠 Hydro Tasmania）","C&I 灵活资产 VPP + 零售",
  "VIC/NSW/QLD/SA 零售商；C&I 灵活资产聚合（MW 未披露）",
  "光伏/电池/冷库/数据中心/HVAC AI 调控 + 售电","零售套餐 + VPP 收益一体化"),
 ("BESS\n优化 / 交易","Hachiko Energy","BESS 优化/组合管理 SaaS","C&I / <5MW 非计划 BESS",
  "2024 成立 · $2.5M 种子轮 · 基本 pre-traction（未披露在管 MW）",
  "电池组合优化、多市场报价（套利+FCAS），宣称收益 +75%","纯 SaaS 软件费，不分成（收益全归客户）"),
 ("BESS\n优化 / 交易","Tesla Autobidder","Autobidder 实时交易平台","大电站级 BESS（utility）",
  "Hornsdale(SA) 等；多国市场领先","电池实时交易 / 组合优化（套利+FCAS+辅助服务）","软件授权（配 Tesla 硬件）"),
 ("BESS\n优化 / 交易","Fluence Mosaic","Mosaic（AI 报价）+ Nispera","大电站级 BESS（utility）",
  "全球 >10GW 在管；Akaysha Waratah 850MW / Orana 415MW、AMPYR 300MW",
  "AI 价格预测 + 报价优化、资产绩效管理","软件订阅 / 服务费（SaaS）"),
 ("DER 编排\n平台（软件）","SwitchDin","Stormcloud(云)+Droplet(控制器)","C&I/户用/微网 DER 编排",
  "PowerPlus BESS；Alice Springs / NT 首个 VPP 等多项目",
  "多品牌设备级接入、VPP/微网编排、DR、辅助服务","软件平台/授权（SaaS + 硬件控制器）"),
 ("DER 编排\n平台（软件）","Evergen","Intelligent Control / Evergen ONE","户用为主 VPP / EMS",
  "编排数千户家庭电池；与 Enphase/LG 集成","卖软件给零售商/电网编排家庭电池、电池延寿","SaaS / 平台授权（B2B 软件费）"),
 ("DER 编排\n平台（软件）","Reposit Power","GridCredits","户用",
  "零售商无关的软件（retailer-agnostic）","家庭电池交易/调度软件，把家庭变“交易员”","软件授权"),
 ("零售商型\nVPP（户用）","Amber Electric","SmartShift / Amber for Batteries","户用（+小工商）现货敞口",
  "户用 VPP 头部（非 Tesla 电池最灵活）","批发现货价透传 + 电池自动优化（套利）","月订阅费（收益归客户）"),
 ("零售商型\nVPP（户用）","AGL / Origin Loop","零售商 VPP","户用为主",
  "约 10 家活跃户用 VPP 之列；支持 BYD/Sungrow/AlphaESS/Tesla",
  "电池 VPP、事件 credit、set-and-forget","零售套餐 + 签约/事件 bill credit"),
 ("白标 / 通道","Localvolts / Energy Locals / Powow","零售/市场接入","C&I/户用 白标",
  "—（既是潜在通道，也是潜在对手）","零售白标、市场结算通道","服务费 / 收益抽成"),
 ("我方\n（对照）","星星充电 StarCharge","自有 VPP 平台 + 微网运营","C&I/<5MW BESS 主轴 + 户用",
  "试点 50MW → 中期 320MW → 远期 650MW（规划）",
  "代运营（套利+FCAS+DR）+ 平台授权 + EV/V2G + 微网","自营分成 30% + 平台授权 15% + BESS 兜底"),
]

r=4
group_start=r; prev=rows[0][0]
def fill_for(cat): return TINT if cat.startswith("我方") else GREY
for i,row in enumerate(rows):
    cat=row[0]
    for j,val in enumerate(row,1):
        cell=ws.cell(row=r,column=j,value=val); cell.border=border
        cell.alignment=ctr if j in (1,2) else wrap
        if j==1:
            cell.font=FN(9.5,True,ORANGE if not cat.startswith("我方") else "B5400F")
        elif j==2:
            cell.font=FN(10,True, "B5400F" if cat.startswith("我方") else INK)
        else:
            cell.font=FN(9.3, cat.startswith("我方"))
        if cat.startswith("我方"):
            cell.fill=PatternFill("solid",fgColor=TINT)
    ws.row_dimensions[r].height=46
    r+=1

# merge 类别 column per consecutive group
def merge_groups():
    start=4; cur=ws.cell(row=4,column=1).value
    for rr in range(5,4+len(rows)+1):
        v=ws.cell(row=rr,column=1).value if rr<4+len(rows) else None
        if v!=cur:
            if rr-1>start:
                ws.merge_cells(start_row=start,start_column=1,end_row=rr-1,end_column=1)
            start=rr; cur=v
merge_groups()
for rr in range(4,4+len(rows)):
    ws.cell(row=rr,column=1).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
ws.freeze_panes="A4"
wb.save(OUT); print("saved",OUT)
